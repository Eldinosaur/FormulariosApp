import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

FILE_PATH = 'data/data.xlsx'


def checkbox_to_text(value):
    """Convierte checkbox a texto legible"""
    if value is True or value == "si" or value == "sí" or value == "true":
        return "Sí"
    elif value is False or value == "no" or value == "false":
        return "No"
    return str(value) if value else ""


def format_money_excel(value):
    """Formatea valores monetarios para Excel (como número, no string)"""
    try:
        num = float(value) if value else 0
        return num
    except:
        return 0


def create_excel_file(data):
    """Crea o actualiza el archivo Excel con todos los datos del formulario"""
    
    # Definir todos los headers (completos)
    HEADERS = [
        # ========== METADATOS ==========
        "Fecha registro",
        "Fecha visita",
        
        # ========== 1. DATOS PERSONALES ==========
        "CI", 
        "Nombre completo", 
        "Edad", 
        "Fecha nacimiento lugar",
        "Género", 
        "Estado civil", 
        "Email", 
        "Teléfono", 
        "Dirección", 
        "Contacto emergencia",
        
        # ========== 2. DATOS LABORALES ==========
        "Área trabajo", 
        "Puesto", 
        "Fecha ingreso",
        
        # ========== 3. EDUCACIÓN ==========
        "Nivel educativo", 
        "Observaciones educación", 
        "Estudia actualmente", 
        "Carrera", 
        "Nivel estudio",
        
        # ========== 4. DISCAPACIDAD ==========
        "Tiene discapacidad", 
        "Tipo discapacidad", 
        "Porcentaje discapacidad",
        
        # ========== 5. VIVIENDA ==========
        "Sector", 
        "Tipo vivienda", 
        "Clase vivienda", 
        "Tenencia", 
        "N° habitantes", 
        "Tiempo sector", 
        "Sector seguro", 
        "Avalúo", 
        "Observaciones vivienda",
        
        # ========== 6. DISTRIBUCIÓN VIVIENDA ==========
        "Dormitorio", 
        "Camas", 
        "Cocina", 
        "Comedor", 
        "Sala", 
        "Baño", 
        "Patio", 
        "Jardín", 
        "Terraza", 
        "Garaje",
        
        # ========== 7. MATERIALES ==========
        "Techo material", 
        "Techo estado", 
        "Techo otros",
        "Pared material", 
        "Pared estado", 
        "Pared otros",
        "Piso material", 
        "Piso estado", 
        "Piso otros",
        "Estructura material", 
        "Estructura estado", 
        "Estructura otros",
        
        # ========== 8. SERVICIOS BÁSICOS ==========
        "Tipo agua", 
        "Tipo luz", 
        "Tipo SSHH", 
        "Observaciones servicios", 
        "Hacinamiento",
        "Manejo desechos", 
        "Tipo transporte", 
        "Transporte otro", 
        "Electrodomésticos",
        "Tiene Internet", 
        "Tipo Internet",
        
        # ========== 9. ANIMALES ==========
        "Tiene animales", 
        "Tipo animales", 
        "Cantidad animales", 
        "Zona peste", 
        "Lugar tenencia", 
        "Observaciones animales",
        
        # ========== 10. INFORMACIÓN FAMILIAR ==========
        "Tipo familia", 
        "N° hijos", 
        "Hijos fuera matrimonio", 
        "Paga pensión",
        "N° matrimonio", 
        "Comparte tiempo familia", 
        "Frecuencia tiempo familiar", 
        "Actividades familiares",
        "Actividades fuera trabajo", 
        "Otras actividades domicilio", 
        "Especificar otras actividades",
        "Hobbies", 
        "Tiempo hobbies", 
        "Relación pareja", 
        "Por qué relación pareja",
        "Relación hijos", 
        "Por qué relación hijos", 
        "Problemas familiares", 
        "Recibió ayuda problemas",
        "Migración familiar", 
        "Recibió dinero exterior",
        
        # ========== 11. SITUACIÓN ECONÓMICA - APORTES ==========
        "Gastos compartidos",
        "Aporte padre", 
        "Aporte madre", 
        "Aporte hermanos", 
        "Aporte colaborador",
        "Aporte cónyuge", 
        "Aporte hijos", 
        "Aporte otros", 
        "Total aportes",
        "Monto deudas", 
        "Préstamos formales", 
        "Monto préstamos formales",
        "Préstamos informales", 
        "Préstamos familiares", 
        "Préstamos chulqueros", 
        "Otros préstamos informales", 
        "Tarjetas crédito",
        
        # ========== 12. GASTOS MENSUALES ==========
        "Gasto alimentación", 
        "Gasto educación", 
        "Gasto vivienda", 
        "Gasto vestimenta",
        "Gasto salud", 
        "Gasto transporte", 
        "Gasto servicios básicos", 
        "Gasto internet",
        "Gasto TV cable", 
        "Gasto plan celular", 
        "Gasto préstamos", 
        "Gasto préstamos quirografarios",
        "Gasto tarjetas crédito", 
        "Gasto pensión alimentos", 
        "Gasto locales comerciales",
        "Gasto apoyo terceros", 
        "Otros gastos", 
        "Total gastos",
        
        # ========== 13. INGRESOS Y BALANCE ==========
        "Total ingresos", 
        "Balance (Ingresos - Gastos)", 
        "Posee vehículo", 
        "Descripción vehículo",
        "Actividad económica adicional", 
        "Crianza animales",
        
        # ========== 14. SALUD ==========
        "Practica deporte", 
        "Deporte que practica", 
        "Frecuencia deporte",
        "Tiene enfermedad", 
        "Problemas salud familia", 
        "Descripción problemas salud",
        "Discapacidad familia", 
        "Tipo discapacidad familia", 
        "Porcentaje discapacidad familia",
        "Parentezco discapacidad",
        
        # ========== 15. SITUACIÓN LABORAL ==========
        "Ocupación anterior", 
        "Funciones actuales", 
        "Relación trabajo-formación",
        "Relación compañeros", 
        "Qué mejoraría", 
        "Resolución conflictos",
        "Trabajo desgastante", 
        "Presión laboral", 
        "Genera estrés", 
        "Tiempo suficiente",
        "Reconocimiento jefe", 
        "Reconocimiento empresa", 
        "Proyección laboral", 
        "Ha sufrido discriminación",
        "Cómo mejorar como trabajador", 
        "Beneficios sugeridos"
    ]
    
    # ========== AGREGAR HEADERS DE FAMILIARES (9 CAMPOS POR MIEMBRO) ==========
    # Campos: Nombres, Edad, Parentezco, Estado Civil, Instrucción, Ocupación, Lugar trabajo/estudio, Ingresos, Teléfono
    for i in range(1, 7):  # Máximo 6 miembros
        HEADERS.extend([
            f"F{i}_Nombres",
            f"F{i}_Edad",
            f"F{i}_Parentezco",
            f"F{i}_EstadoCivil",
            f"F{i}_Instruccion",
            f"F{i}_Ocupacion",
            f"F{i}_LugarTrabajoEstudio",
            f"F{i}_Ingresos",
            f"F{i}_Telefono"
        ])
    
    # Crear archivo si no existe
    if not os.path.exists(FILE_PATH):
        os.makedirs('data', exist_ok=True)
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Fichas Socioeconómicas"
        
        # Estilos para headers
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Escribir headers con estilos
        for col_idx, header in enumerate(HEADERS, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Ajustar anchos de columnas
        for col_idx, header in enumerate(HEADERS, 1):
            column_letter = worksheet.cell(row=1, column=col_idx).column_letter
            worksheet.column_dimensions[column_letter].width = min(max(len(str(header)) + 3, 12), 30)
        
        workbook.save(FILE_PATH)
    
    # Cargar archivo existente
    workbook = load_workbook(FILE_PATH)
    worksheet = workbook.active
    
    # ========== CONSTRUIR FILA CON TODOS LOS DATOS ==========
    row = []
    
    # METADATOS
    row.append(datetime.now().strftime("%d/%m/%Y %H:%M"))
    row.append(data.get("visit_date", ""))
    
    # 1. DATOS PERSONALES
    row.append(data.get("ci", ""))
    row.append(data.get("name", ""))
    row.append(data.get("age", ""))
    row.append(data.get("birth_date_place", ""))
    row.append(data.get("gender", ""))
    row.append(data.get("marital_status", ""))
    row.append(data.get("email", ""))
    row.append(data.get("phone", ""))
    row.append(data.get("address", ""))
    row.append(data.get("emergency_contact", ""))
    
    # 2. DATOS LABORALES
    row.append(data.get("work_area", ""))
    row.append(data.get("job_title", ""))
    row.append(data.get("entry_date", ""))
    
    # 3. EDUCACIÓN
    row.append(data.get("educational_level", ""))
    row.append(data.get("educational_observations", ""))
    row.append(checkbox_to_text(data.get("currently_studying")))
    row.append(data.get("study_career", ""))
    row.append(data.get("study_level", ""))
    
    # 4. DISCAPACIDAD
    row.append(checkbox_to_text(data.get("disability")))
    row.append(data.get("disability_type", ""))
    row.append(data.get("disability_percentage", ""))
    
    # 5. VIVIENDA
    row.append(data.get("sector", ""))
    row.append(data.get("house_type", ""))
    row.append(data.get("house_class", ""))
    row.append(data.get("house_ownership", ""))
    row.append(data.get("household_size", ""))
    row.append(data.get("time_living_sector", ""))
    row.append(checkbox_to_text(data.get("is_safe")))
    row.append(format_money_excel(data.get("house_valuation")))
    row.append(data.get("house_observations", ""))
    
    # 6. DISTRIBUCIÓN VIVIENDA
    distribution = data.get("house_distribution", {})
    distribution_keys = ["dormitorio", "camas", "cocina", "comedor", "sala", 
                         "baño", "patio", "jardín", "terraza", "garaje"]
    for key in distribution_keys:
        value = distribution.get(key, 0)
        row.append(value if value > 0 else "No")
    
    # 7. MATERIALES
    row.append(data.get("roof_type", ""))
    row.append(data.get("roof_status", ""))
    row.append(data.get("roof_type_other", ""))
    row.append(data.get("wall_type", ""))
    row.append(data.get("wall_status", ""))
    row.append(data.get("wall_type_other", ""))
    row.append(data.get("floor_type", ""))
    row.append(data.get("floor_status", ""))
    row.append(data.get("floor_type_other", ""))
    row.append(data.get("structure_type", ""))
    row.append(data.get("structure_status", ""))
    row.append(data.get("structure_type_other", ""))
    
    # 8. SERVICIOS BÁSICOS
    row.append(data.get("water_type", ""))
    row.append(data.get("light_type", ""))
    row.append(data.get("sshh_type", ""))
    row.append(data.get("basic_services_other", ""))
    row.append(checkbox_to_text(data.get("hacinamiento")))
    row.append(data.get("waste_management", ""))
    row.append(data.get("transport_type", ""))
    row.append(data.get("transport_type_other", ""))
    row.append(data.get("appliances", ""))
    row.append(checkbox_to_text(data.get("internet")))
    row.append(data.get("internet_type", ""))
    
    # 9. ANIMALES
    row.append(checkbox_to_text(data.get("animals")))
    row.append(data.get("animal_type", ""))
    row.append(data.get("animal_quantity", ""))
    row.append(checkbox_to_text(data.get("peste")))
    row.append(data.get("animal_location", ""))
    row.append(data.get("animal_observations", ""))
    
    # 10. INFORMACIÓN FAMILIAR
    row.append(data.get("family_type", ""))
    row.append(data.get("children_count", ""))
    row.append(checkbox_to_text(data.get("children_outside_marriage")))
    row.append(checkbox_to_text(data.get("pays_alimony")))
    row.append(data.get("marriage_number", ""))
    row.append(checkbox_to_text(data.get("family_time")))
    row.append(data.get("family_time_frequency", ""))
    row.append(data.get("family_activities", ""))
    row.append(data.get("other_activities", ""))
    # Para "Otras actividades domicilio" (checkbox)
    other_activities_val = data.get("other_activities")
    if isinstance(other_activities_val, bool) or other_activities_val in ["si", "no"]:
        row.append(checkbox_to_text(other_activities_val))
    else:
        row.append(other_activities_val if other_activities_val else "")
    row.append(data.get("other_activities_specify", ""))
    row.append(data.get("hobbies", ""))
    row.append(data.get("hobbies_time", ""))
    row.append(data.get("partner_relationship", ""))
    row.append(data.get("partner_relationship_reason", ""))
    row.append(data.get("children_relationship", ""))
    row.append(data.get("children_relationship_reason", ""))
    row.append(data.get("family_problems", ""))
    row.append(checkbox_to_text(data.get("family_problems_help")))
    row.append(checkbox_to_text(data.get("family_migration")))
    row.append(checkbox_to_text(data.get("family_migration_received")))
    
    # 11. SITUACIÓN ECONÓMICA - APORTES
    row.append(checkbox_to_text(data.get("shared_expenses")))
    row.append(format_money_excel(data.get("father_contribution")))
    row.append(format_money_excel(data.get("mother_contribution")))
    row.append(format_money_excel(data.get("siblings_contribution")))
    row.append(format_money_excel(data.get("collaborators_contribution")))
    row.append(format_money_excel(data.get("spouse_contribution")))
    row.append(format_money_excel(data.get("children_contribution")))
    row.append(format_money_excel(data.get("other_contribution")))
    row.append(format_money_excel(data.get("total_contribution")))
    row.append(format_money_excel(data.get("debt_amount")))
    row.append(checkbox_to_text(data.get("formal_loans")))
    row.append(format_money_excel(data.get("formal_loans_amount")))
    row.append(checkbox_to_text(data.get("informal_loans")))
    row.append(format_money_excel(data.get("informal_loans_family_amount")))
    row.append(format_money_excel(data.get("informal_loans_moneylender_amount")))
    row.append(format_money_excel(data.get("informal_loans_other_amount")))
    row.append(checkbox_to_text(data.get("credit_cards")))
    
    # 12. GASTOS MENSUALES
    row.append(format_money_excel(data.get("food_support")))
    row.append(format_money_excel(data.get("education_support")))
    row.append(format_money_excel(data.get("housing_support")))
    row.append(format_money_excel(data.get("clothing_support")))
    row.append(format_money_excel(data.get("health_support")))
    row.append(format_money_excel(data.get("transport_support")))
    row.append(format_money_excel(data.get("basic_services_support")))
    row.append(format_money_excel(data.get("internet_support")))
    row.append(format_money_excel(data.get("cable_tv_support")))
    row.append(format_money_excel(data.get("cell_plan_support")))
    row.append(format_money_excel(data.get("loans_support")))
    row.append(format_money_excel(data.get("unsecured_loans_support")))
    row.append(format_money_excel(data.get("credit_cards_support")))
    row.append(format_money_excel(data.get("alimony_support")))
    row.append(format_money_excel(data.get("commercial_properties_support")))
    row.append(format_money_excel(data.get("financial_support_others")))
    row.append(format_money_excel(data.get("other_expenses_support")))
    row.append(format_money_excel(data.get("total_expenses")))
    
    # 13. INGRESOS Y BALANCE
    row.append(format_money_excel(data.get("total_income")))
    row.append(format_money_excel(data.get("balance")))
    row.append(checkbox_to_text(data.get("transportation")))
    row.append(data.get("transportation_description", ""))
    row.append(data.get("additional_economic_activity", ""))
    row.append(checkbox_to_text(data.get("animal_breeding")))
    
    # 14. SALUD
    row.append(checkbox_to_text(data.get("sports")))
    row.append(data.get("sports_description", ""))
    row.append(data.get("sports_frequency", ""))
    row.append(checkbox_to_text(data.get("disease")))
    row.append(checkbox_to_text(data.get("family_health_problems")))
    row.append(data.get("family_health_problems_description", ""))
    row.append(checkbox_to_text(data.get("family_disability")))
    row.append(data.get("family_disability_type", ""))
    row.append(data.get("family_disability_percentage", ""))
    row.append(data.get("family_disability_relationship", ""))
    
    # 15. SITUACIÓN LABORAL
    row.append(data.get("previous_occupation", ""))
    row.append(data.get("current_functions", ""))
    row.append(data.get("job_relation", ""))
    row.append(data.get("colleague_relationship", ""))
    row.append(data.get("improvement_suggestions", ""))
    row.append(data.get("conflict_resolution", ""))
    row.append(data.get("job_exhaustion", ""))
    row.append(data.get("job_pressure", ""))
    row.append(data.get("job_stress", ""))
    row.append(data.get("job_time_management", ""))
    row.append(data.get("job_manager_recognition", ""))
    row.append(data.get("job_recognition", ""))
    row.append(data.get("job_projection", ""))
    row.append(checkbox_to_text(data.get("job_discrimination")))
    row.append(data.get("job_improvement", ""))
    row.append(data.get("job_benefits", ""))
    
    # ========== FAMILIARES (MÁXIMO 6 MIEMBROS - 9 CAMPOS CADA UNO) ==========
    family_members = data.get("family_members", [])
    
    for i in range(6):  # Máximo 6 miembros
        if i < len(family_members):
            member = family_members[i]
            row.extend([
                member.get("name", ""),
                member.get("age", ""),
                member.get("relation", ""),
                member.get("marital_status", ""),
                member.get("education", ""),
                member.get("job", ""),
                member.get("work_or_study_place", ""),
                format_money_excel(member.get("income", 0)),
                member.get("phone", "")
            ])
        else:
            # Campos vacíos si no hay suficientes familiares
            row.extend(["", "", "", "", "", "", "", "", ""])
    
    # Verificar que la fila tenga la longitud correcta
    expected_length = len(HEADERS)
    if len(row) != expected_length:
        print(f"Advertencia: Longitud de fila ({len(row)}) no coincide con headers ({expected_length})")
        if len(row) < expected_length:
            row.extend([""] * (expected_length - len(row)))
        else:
            row = row[:expected_length]
    
    # Agregar la fila al worksheet
    worksheet.append(row)
    
    # Autoajustar anchos de columnas
    for col_idx in range(1, len(HEADERS) + 1):
        column_letter = worksheet.cell(row=1, column=col_idx).column_letter
        max_length = len(str(HEADERS[col_idx - 1]))
        
        for row_idx in range(2, min(worksheet.max_row, 20)):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        worksheet.column_dimensions[column_letter].width = min(max_length + 3, 40)
    
    # Guardar el archivo
    workbook.save(FILE_PATH)
    
    return FILE_PATH


def get_all_records():
    """Obtiene todos los registros del archivo Excel"""
    if not os.path.exists(FILE_PATH):
        return []
    
    workbook = load_workbook(FILE_PATH)
    worksheet = workbook.active
    
    records = []
    headers = [cell.value for cell in worksheet[1]]
    
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if any(row):
            record = dict(zip(headers, row))
            records.append(record)
    
    return records


def get_record_by_ci(ci):
    """Busca un registro por cédula"""
    records = get_all_records()
    for record in records:
        if str(record.get("CI", "")) == str(ci):
            return record
    return None