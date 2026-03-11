import os
from openpyxl import Workbook, load_workbook

# Ruta del archivo Excel
FILE_PATH = 'data/data.xlsx'

# Función para crear el archivo Excel y agregar datos
def create_excel_file(data):
    # Verificar si el archivo ya existe, si no, crear uno nuevo
    if not os.path.exists(FILE_PATH):
        # Crear un nuevo libro de trabajo
        workbook = Workbook()
        worksheet = workbook.active
        # Agregar encabezados a la hoja de cálculo
        worksheet.append(['ID', 'Name', 'Email'])
        workbook.save(FILE_PATH)
        
    # Cargar el libro de trabajo existente    
    workbook = load_workbook(FILE_PATH)
    worksheet = workbook.active
    # Agregar los datos a la hoja de cálculo
    worksheet.append([
        data['id'],
        data['name'],
        data['email']
    ])
    workbook.save(FILE_PATH)